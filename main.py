from flask import Flask, request, render_template, redirect, url_for
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from io import BytesIO
import logging
from flask_session import Session
import re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

app.config['SESSION_TYPE'] = 'filesystem'
Session(app)



@app.route('/', methods=['GET'])
def registration_form():
    areas_with_schools = read_excel_data()
    return render_template('index.html', areas_with_schools=areas_with_schools)

def read_excel_data():
    workbook = load_workbook("Schools.xlsx")
    sheet = workbook.active
    areas_with_schools = {}

    schools_to_remove = [
        '"Құрама бастауыш мектебі" КММ мектептің жабылуына құжаттар дайындалуда',
        '"Андрейниковка бастауыш мектебі" КММ мектептің жабылуына құжаттар дайындалуда',
        'Уақытша жабылды "АЗИЯ" комплекс-мектебі м-сі',
        '"Философиялық бағыттағы жалпы білім беретін жеке мектеп" ЖММ (уақытша жұмысы тоқтатылды)',
        '"№31 жалпы орта білім беретін мектебі" КММ (жұмысы тоқтатылды)',
        '"№44 негізгі орта мектебі" КММ (қызметі уақытша тоқтатылды)',
        '"Ассоциированная школа ЮНЕСКО "ДИАЛОГ" ЖММ (жұмысы тоқтатылған)',
        '(Уақытша жұмысы тоқтатылған) Шығыс Қазақстан облысы білім басқармасы Күршім ауданы бойынша білім бөлімінің «Топтерек бастауыш мектебі» коммуналдық мемлекеттік мекемесі',
        '(Уақытша тоқтатылды) Шығыс Қазақстан облысы білім басқармасы Күршім ауданы бойынша білім бөлімінің «Жиделі бастауыш мектебі» коммуналдық мемлекеттік мекемесі',
        '"Math-Language "SEED SCHOOL" ЖМ (тоқтатылған)',
        'КЕБМ"Экономикалық лицей" (тоқтатылған)',
        '"Қазақ ұлттық хореография академиясы" мектебі ҚЕАҚ (Тоқтатылды)',
        '"MLS ELORDA" ЖШС (Тоқтатылған)',
        ' «National Academy of Education «ULAGAT» мектебі (Тоқтатылған)',
        '-',
        '"Көкшетау қаласындағы ФМБ НЗМ"',
        'Атырау қаласындағы "ХББ НЗМ" филиалы',
        '"Ақтөбе қаласындағы ФМБ НЗМ" филиалы ',
        '"Орал қаласындағы ФМБ НЗМ" филиалы ',
        '"Талдықорған қаласындағы ФМБ НЗМ" филиалы',
        'Қарағанды қаласындағы "ХББ НЗМ" ф-лы',
        '"Қостанай қаласындағы ФМБ НЗМ"',
        '"Тараз қаласындағы ФМБ НЗМ" филиалы',
        'Қызылорда қаласындағы ХББ НЗМ филиалы',
        'Ақтау қаласындағы ХББ НЗМ филиалы',
        'Шымкент қ. "химия-биологиялық бағыттағы НЗМ" филиалы',
        'Шымкент қ. "физика-математикалық бағыттағы НЗМ" филиалы',
        'Павлодар қаласындағы ХББ НЗМ филиалы',
        ' "Өскемен қаласының ХББ НЗМ"  филиалы',
        '"Семей қаласындағы ФМБ НЗМ"',
        '"Петропавл қаласының ХББ НЗМ" филиалы',
        '"НЗМ"  ДБҰ "Астана қаласындағы Назарбаев Зияткерлік мектебі" филиалы',
        'Астана қаласындағы "физика-математикалық бағытындағы Назарбаев Зияткерлік мектебі" "НЗМ"  ДБҰ филиалы',
        '"НЗМ" ДБҰ "Астана қаласындағы Халықаралық мектеп" филиалы',
        '"Алматы қаласындағы ХББ НЗМ" филиалы',
        '"Алматы қаласындағы ФМБ НЗМ" филиалы ',
        'Түркістан қ. "химия-биологиялық бағыттағы НЗМ" филиалы',
        '"Республикалық физика-математикалық мектебі" КЕАҚ филиалы',
        '"Республикалық физика-математика мектебі" КЕАҚ филиалы',
        'Орал қ. «Республикалық физика-математикалық мектебі» Коммерциалық емес акционерлік қоғамының филиалы',
        '"Абай атындағы Республикалық мамандандырылған дарынды балаларға арналған қазақ тілі мен әдебиетін тереңдете оқытатын орта мектеп- интернаты" РММ'

    ]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        area = row[4]  # Assuming area is in the first column (0-indexed)
        school = row[6]  # Assuming schools are in the second column (0-indexed)

        if school in schools_to_remove:
            continue

        if area in areas_with_schools:
            areas_with_schools[area].append(school)
        else:
            areas_with_schools[area] = [school]

    areas_with_schools["«Бөбек» Ұлттық ғылыми-практикалық, білім беру және сауықтыру орталығы» РМҚК"] = ["«Бөбек» Ұлттық ғылыми-практикалық, білім беру және сауықтыру орталығы» РМҚК"]
    areas_with_schools["НЗМ ДББҰ"] = ["Көкшетау қаласындағы ФМБ НЗМ", "Атырау қаласындағы ХББ НЗМ филиалы", "Ақтөбе қаласындағы ФМБ НЗМ филиалы",
                                      "Орал қаласындағы ФМБ НЗМ филиалы", "Талдықорған қаласындағы ФМБ НЗМ филиалы", "Қарағанды қаласындағы ХББ НЗМ филиалы",
                                      "Қостанай қаласындағы ФМБ НЗМ", "Тараз қаласындағы ФМБ НЗМ филиалы", "Қызылорда қаласындағы ХББ НЗМ филиалы",
                                      "Ақтау қаласындағы ХББ НЗМ филиалы", "Шымкент қ. химия-биологиялық бағыттағы НЗМ филиалы", "Шымкент қ. физика-математикалық бағыттағы НЗМ филиалы",
                                      "Өскемен қаласының ХББ НЗМ  филиалы", "Семей қаласындағы ФМБ НЗМ", "Петропавл қаласының ХББ НЗМ филиалы", "НЗМ ДБҰ Астана қаласындағы Назарбаев Зияткерлік мектебі филиалы",
                                      "Астана қаласындағы физика-математикалық бағытындағы Назарбаев Зияткерлік мектебі НЗМ  ДБҰ филиалы", "НЗМ ДБҰ Астана қаласындағы Халықаралық мектеп филиалы",
                                      "Алматы қаласындағы ХББ НЗМ филиалы", "Алматы қаласындағы ФМБ НЗМ филиалы", "Түркістан қ. химия-биологиялық бағыттағы НЗМ филиалы", "Павлодар қаласындағы ХББ НЗМ филиалы"]
    areas_with_schools["РФММ КеАҚ"] = [
        "Астана қ. Республикалық физика-математикалық мектебі КЕАҚ филиалы", "Алматы қ. Республикалық физика-математикалық мектебі КЕАҚ филиалы", "Орал қ. Республикалық физика-математикалық мектебі КЕАҚ филиалы"]
    areas_with_schools["Абай атындағы РММИ"] = ["Абай атындағы Республикалық мамандандырылған дарынды балаларға арналған қазақ тілі мен әдебиетін тереңдете оқытатын орта мектеп- интернаты РММ"]
    return areas_with_schools


@app.route('/submit', methods=['POST'])
def submit_registration():
    try:
        # Get form data (same as before)
        field_data = {
            'Аймақтың атауы': request.form['area'],
            'Мектебі': request.form['school'],
            'Қатысушының ЖСНі': request.form['participant_iin'],
            'Қатысушының аты-жөні': request.form['participant_name'],
            'Туған күні': request.form['birth_date'],
            'Қатысушының жынысы': request.form['participant_gender'],
            'Топтық / жеке': request.form['group/individual'],
            'Сыныбы': request.form['participant_class'],
            'Қалалық / Ауылдық': request.form['city/rural'],
            'Оқу тілі': request.form['language'],
            'Секциясы / Секция': request.form['section'],
            'Тақырыбы / Тема': request.form['project_title'],
            '1-ші жетекшінің аты-жөні': request.form['1st_supervisor_name'],
            '1-ші жетекшінің ЖСНі': request.form['1st_supervisor_iin'],
            '2-ші жетекшінің аты-жөні': request.form['2nd_supervisor_name'],
            '2-ші жетекшінің ЖСНі': request.form['2nd_supervisor_iin']
        }

        # Validate participant IIN matches birth date
        iin = request.form['participant_iin']
        birth_date = request.form['birth_date'] 
        if len(iin) >= 6 and birth_date:
            year, month, day = birth_date.split('-')
            expected_prefix = year[-2:] + month + day
            if iin[:6] != expected_prefix:
                return "Ошибка: Первые 6 цифр ИИН должны совпадать с датой рождения (ГГММДД)", 400
        
        for sup_key in ['1st_supervisor_iin', '2nd_supervisor_iin']:
            sup_iin = request.form.get(sup_key, '')
            if sup_iin and not re.fullmatch(r'\d{12}', sup_iin):
                return f"Ошибка: ИИН {sup_key} должен содержать ровно 12 цифр", 400

        # Validate names: only letters, each word capitalized, no digits or symbols
        name_fields = ['participant_name', '1st_supervisor_name', '2nd_supervisor_name']
        name_pattern = re.compile(r"^[A-Za-zА-Яа-яӘәҒғҚқҢңӨөҰұҮүHhІіЁёЫыІіЭэҮүҰұҚқҒғӘәӨөҺһ\s'-]+$")
        for field in name_fields:
            name = request.form.get(field, '')
            if not name_pattern.fullmatch(name):
                return f"Ошибка: '{field}' должно содержать только буквы и пробелы, без цифр и символов", 400
            words = [w for w in name.split(' ') if w]
            for word in words:
                if word[0] != word[0].upper():
                    return f"Ошибка: В поле '{field}' каждое слово должно начинаться с заглавной буквы", 400

        participant_values = list(field_data.values())

        uploaded_file = request.files['file']
        if uploaded_file.filename == "":
            return "Ошибка: Файл не выбран.", 400

        # Save file locally
        os.makedirs("uploads", exist_ok=True)
        file_path = os.path.join("uploads", uploaded_file.filename)
        uploaded_file.save(file_path)

        # Save data to Excel
        excel_file = 'registrations.xlsx'
        if not os.path.exists(excel_file):
            # Create file if not exists
            wb = Workbook()
            ws = wb.active
            ws.append(list(field_data.keys()))  # header
        else:
            wb = load_workbook(excel_file)
            ws = wb.active

        ws.append(participant_values)
        wb.save(excel_file)

        return redirect(url_for('registration_form'))

    except Exception as e:
        print(f"Произошла ошибка: {str(e)}")
        return "Внутренняя ошибка сервера", 500

logging.basicConfig(level=logging.INFO, filename='app.log', filemode='a', format='%(name)s - %(levelname)s - %(message)s')
logging.info("Приложение запущено")
logging.error("Произошла ошибка при обработке запроса")


def validate_iin(iin):
    """ Validate Kazakhstani ID format. """
    return re.match(r'^\d{12}$', iin)

def validate_email(email):
    """ Validate email format. """
    return re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email)

def send_email(recipient, subject, body):
    """ Send email. """
    msg = MIMEMultipart()
    msg['From'] = 'MYEMAIL'
    msg['To'] = recipient
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))
    server = smtplib.SMTP('smtp.example.com', 587)
    server.starttls()
    server.login(msg['From'], 'your password')
    server.send_message(msg)
    server.quit()

if __name__ == '__main__':
    app.run(debug=True, port=8080)  # SSL for HTTPS
